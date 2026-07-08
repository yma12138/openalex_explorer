#!/usr/bin/env python3
"""项目管理 — 列出 / 查看 / 删除项目，清理密钥 | Project Management.

用法 | Usage:
  python manage.py list                          # 列出所有项目 | List all projects
  python manage.py show <index|path>             # 查看项目详情 | Show project details
  python manage.py delete <index|path>           # 删除项目（含确认）| Delete a project (with confirm)
   python manage.py clean-secrets                 # 清理 API key | Clean API keys from configs
   python manage.py export <index|path>           # 导出论文到 Excel | Export papers to Excel
"""

import argparse
import json
import shutil
import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
LLM_CONFIG = Path(__file__).resolve().parent / "llm_config.json"


# ── 共享工具 ────────────────────────────────────────────


def _db_stats(db_path: Path) -> dict:
    """Query paper_store.db and return stats dict."""
    if not db_path.exists():
        return {}
    try:
        conn = sqlite3.connect(str(db_path))
        cur = conn.execute("SELECT status, COUNT(*) FROM papers GROUP BY status")
        statuses = {row[0]: row[1] for row in cur.fetchall()}
        total = sum(statuses.values()) if statuses else 0
        conn.close()
        return {
            "total": total,
            "searched": statuses.get("searched", 0),
            "relevanced": statuses.get("relevanced", 0),
            "quality_passed": statuses.get("quality_passed", 0),
            "rejected": statuses.get("rejected", 0),
            "summarized": statuses.get("summarized", 0),
        }
    except (sqlite3.DatabaseError, OSError):
        return {}


def _has_review(dir_path: Path) -> bool:
    return (dir_path / "review.md").exists()


def _parse_session_name(name: str) -> tuple[str, str]:
    """Extract (question, timestamp) from '{question}_{YYYYMMDD_HHMMSS}'."""
    if "_202" in name:
        parts = name.rsplit("_202", 1)
        question = parts[0].replace("_", " ")
        ts = "202" + parts[1] if len(parts) > 1 else ""
        if len(ts) >= 15:
            ts = f"{ts[:4]}-{ts[4:6]}-{ts[6:8]} {ts[9:11]}:{ts[11:13]}:{ts[13:15]}"
        return question, ts
    return name.replace("_", " "), ""


def _list_sessions(base_dir: Path) -> list[Path]:
    """Return all session dirs under base_dir (sorted newest first)."""
    if not base_dir.exists():
        return []
    dirs = [d for d in base_dir.iterdir() if d.is_dir()]
    dirs.sort(key=lambda d: d.stat().st_mtime, reverse=True)
    return dirs


def _resolve_session(session_arg: str) -> Path:
    """解析序号或路径 | Resolve index or path to session dir."""
    try:
        idx = int(session_arg)
        sessions = _list_sessions(OUTPUT_DIR)
        if 1 <= idx <= len(sessions):
            return sessions[idx - 1]
        print(f"❌ Index out of range: {idx} (1-{len(sessions)})")
        sys.exit(1)
    except ValueError:
        p = Path(session_arg)
        if p.exists():
            return p if p.is_dir() else p.parent
        print(f"❌ Session not found: {session_arg}")
        sys.exit(1)


# ── Commands ───────────────────────────────────────────


def cmd_list(args):
    """列出所有项目 | List all projects."""
    sessions = _list_sessions(OUTPUT_DIR)
    if not sessions:
        print("No sessions found in output/")
        return

    print(f"{'#':>3}  {'Question':<50} {'Date':<19} {'Papers':>6} {'Passed':>6} {'Review':>6}")
    print("-" * 98)
    for i, s in enumerate(sessions, 1):
        q, ts = _parse_session_name(s.name)
        stats = _db_stats(s / "papers.db")
        has_rev = "✅" if _has_review(s) else ""
        print(
            f"{i:>3}  {q[:48]:<50} {ts[:17]:<19} "
            f"{stats.get('total', '?'):>6} {stats.get('quality_passed', '?'):>6} {has_rev:>6}"
        )
    print("-" * 98)
    print(f"  {len(sessions)} session(s)")


def cmd_show(args):
    """查看项目详情 | Show project details."""
    session_dir = _resolve_session(args.session)
    q, ts = _parse_session_name(session_dir.name)

    print(f"\n{'=' * 60}")
    print(f"  Project: {q}")
    print(f"{'=' * 60}")
    print(f"  Path:  {session_dir}")
    print(f"  Created: {ts}")

    db_path = session_dir / "papers.db"
    stats = _db_stats(db_path)
    if stats:
        print(f"\n  ── Papers ──")
        labels = [("Total searched", "total"),
                  ("Relevanced", "relevanced"),
                  ("Quality passed", "quality_passed"),
                  ("Summarized", "summarized"),
                  ("Rejected", "rejected")]
        for k, v in labels:
            print(f"  {k:28s} {stats.get(v, 0)}")

        if stats.get("quality_passed", 0) > 0:
            try:
                conn = sqlite3.connect(str(db_path))
                cur = conn.execute(
                    "SELECT title, year, quality_score FROM papers "
                    "WHERE status = 'quality_passed' ORDER BY quality_score DESC LIMIT 5"
                )
                papers = cur.fetchall()
                conn.close()
                if papers:
                    print(f"\n  ── Top quality_passed ──")
                    for t, y, s in papers:
                        print(f"    [{s:.0f}] {t[:60]} ({y})")
            except sqlite3.DatabaseError:
                pass

    session_json = session_dir / "session.json"
    if session_json.exists():
        with open(session_json) as f:
            s_info = json.load(f)
        print(f"\n  ── Token Usage ──")
        build_u = s_info.get("build_usage", {})
        if build_u:
            print(f"  Build:    {build_u['total_tokens']:>8,} tokens  ({build_u['calls']} calls)")
        review_u = s_info.get("review_usage", {})
        if review_u:
            print(f"  Review:   {review_u['total_tokens']:>8,} tokens  ({review_u['calls']} calls)")
        total_t = build_u.get("total_tokens", 0) + review_u.get("total_tokens", 0)
        if total_t:
            print(f"  Total:    {total_t:>8,} tokens")

    if _has_review(session_dir):
        rev_path = session_dir / "review.md"
        size = rev_path.stat().st_size
        with open(rev_path, encoding="utf-8") as f:
            lines = f.readlines()
        word_count = sum(len(l.strip().split()) for l in lines if l.strip())
        print(f"\n  ── Review ──")
        print(f"  Path:  {rev_path}")
        print(f"  Size:  {size:,} bytes / ~{word_count} words")
        print(f"  Lines: {len(lines)}")
    else:
        print(f"\n  Review:  ❌ Not yet written")
        print(f"  Run:     python write_review.py {session_dir}")
    print()


def cmd_delete(args):
    """删除项目 | Delete a project."""
    session_dir = _resolve_session(args.session)
    q, _ = _parse_session_name(session_dir.name)

    stats = _db_stats(session_dir / "papers.db")
    paper_count = stats.get("total", "?")
    has_rev = "yes" if _has_review(session_dir) else "no"

    print(f"Project: {q}")
    print(f"  Path:   {session_dir}")
    print(f"  Papers: {paper_count}")
    print(f"  Review: {has_rev}")

    confirm = input(f"\nDelete this project? [y/N] ").strip().lower()
    if confirm != "y":
        print("Cancelled.")
        return

    shutil.rmtree(str(session_dir))
    print(f"✅ Deleted: {session_dir}")


def cmd_clean_secrets(args):
    """清理 API key | Clean API keys."""
    cleaned = 0

    if LLM_CONFIG.exists():
        with open(LLM_CONFIG) as f:
            data = json.load(f)
        if data.get("api_key") and data["api_key"] != "":
            data["api_key"] = ""
            with open(LLM_CONFIG, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            print(f"✅ Cleaned: api_key in {LLM_CONFIG}")
            cleaned += 1
        else:
            print(f"  ℹ️  No api_key in {LLM_CONFIG}")

    for sess in _list_sessions(OUTPUT_DIR):
        sjson = sess / "session.json"
        if not sjson.exists():
            continue
        with open(sjson) as f:
            data = json.load(f)
        modified = False
        if "api_key" in data:
            del data["api_key"]
            modified = True
        if modified:
            with open(sjson, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            print(f"✅ Cleaned: {sjson}")
            cleaned += 1

    openalex_path = Path(__file__).resolve().parent / "openalex_config.json"
    if openalex_path.exists():
        with open(openalex_path) as f:
            data = json.load(f)
        modified = False
        for key in ("mailto", "api_key"):
            if data.get(key, ""):
                data[key] = ""
                modified = True
        if modified:
            with open(openalex_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            print(f"✅ Cleaned: openalex_config.json")
            cleaned += 1

    if cleaned == 0:
        print("  No secrets found.")
    else:
        print(f"\n✅ {cleaned} file(s) cleaned.")


def cmd_export(args):
    """导出论文到 Excel | Export papers to Excel."""
    session_dir = _resolve_session(args.session)
    db_path = session_dir / "papers.db"
    if not db_path.exists():
        print(f"❌ DB not found: {db_path}")
        sys.exit(1)

    q, _ = _parse_session_name(session_dir.name)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("❌ 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    conn = sqlite3.connect(str(db_path))

    # 查所有论文，按 status 排序
    rows = conn.execute(
        "SELECT id, title, authors, year, source, doi, keywords, journal_ref, "
        "status, relevance_score, relevance_reason, quality_score, quality_reason, "
        "summary_json, created_at "
        "FROM papers ORDER BY status, quality_score DESC"
    ).fetchall()

    columns = [desc[0] for desc in conn.execute("PRAGMA table_info(papers)").fetchall()]
    conn.close()

    if not rows:
        print("  DB is empty.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Papers"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Headers (中文 + English)
    headers = [
        "ID", "Title", "Authors", "Year", "Source", "DOI",
        "Keywords", "Journal", "Status", "Relevance", "Relevance Reason",
        "Quality", "Quality Reason", "Summary", "Created",
    ]
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    widths = [12, 60, 40, 8, 12, 30, 30, 20, 14, 10, 30, 10, 30, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Data
    status_colors = {
        "quality_passed": "C6EFCE",
        "rejected": "FFC7CE",
        "summarized": "C6EFCE",
        "relevanced": "BDD7EE",
        "searched": "FCE4D6",
    }

    for r_idx, row in enumerate(rows, 2):
        row_id, title, authors, year, source, doi, keywords, journal_ref = row[:8]
        status, rel_score, rel_reason, qual_score, qual_reason, summary_json, created = row[8:]

        # Parse summary JSON
        summary_text = ""
        if summary_json:
            try:
                sj = json.loads(summary_json)
                parts = []
                for k in ("research_question", "method", "main_findings", "limitations", "relevance_to_topic"):
                    v = sj.get(k, "")
                    if isinstance(v, list):
                        v = "; ".join(v[:3])
                    if v:
                        parts.append(f"{k}: {str(v)[:200]}")
                summary_text = "\n".join(parts)
            except (json.JSONDecodeError, TypeError):
                summary_text = str(summary_json)[:300]

        values = [
            row_id, title, authors, year, source, doi,
            keywords, journal_ref,
            status,
            rel_score, rel_reason or "",
            qual_score, qual_reason or "",
            summary_text,
            created,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            if col_idx == 1:  # ID column
                cell.alignment = Alignment(horizontal="center")

        # Row coloring by status
        color = status_colors.get(status)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=col_idx).fill = fill

    # Freeze header
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"

    # Save
    output_path = session_dir / "papers.xlsx"
    wb.save(str(output_path))
    print(f"\n✅ Exported: {output_path}")
    print(f"   Papers: {len(rows)}")
    print(f"   Open: xdg-open {output_path}")


# ── CLI ─────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="项目管理 | Project Management \n\n"
                    "list / show / delete / clean-secrets"
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_list = sub.add_parser("list", help="列出所有项目 | List all projects")
    p_list.set_defaults(func=cmd_list)

    p_show = sub.add_parser("show", help="查看项目详情 | Show project details")
    p_show.add_argument("session",
                        help="Session 目录路径或 list 中的序号 | Path or index from list")
    p_show.set_defaults(func=cmd_show)

    p_del = sub.add_parser("delete", help="删除项目（需确认）| Delete a project (with confirm)")
    p_del.add_argument("session",
                       help="Session 目录路径或 list 中的序号 | Path or index from list")
    p_del.set_defaults(func=cmd_delete)

    p_clean = sub.add_parser("clean-secrets",
                             help="清理 API key 等敏感信息 | Clean API keys from configs")
    p_clean.set_defaults(func=cmd_clean_secrets)

    p_export = sub.add_parser("export", help="导出论文到 Excel | Export papers to Excel")
    p_export.add_argument("session",
                          help="Session 目录路径或 list 中的序号 | Path or index from list")
    p_export.set_defaults(func=cmd_export)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
